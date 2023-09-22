#!/usr/bin/env python
# coding: utf-8

# In[2]:


import tkinter as tk
import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
import io
from PIL import Image, ImageTk
import tkinter as tk
import tkinter.messagebox
from sklearn.metrics import r2_score
import matplotlib.pyplot as plt
import warnings
from tkinter.filedialog import askopenfilename
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.backend_bases import key_press_handler
from matplotlib.figure import Figure
import sklearn.utils._typedefs
def Addsheet(dfx,name,iox):#增加一个工作表，表名称        
    book = load_workbook(iox)
    writer=pd.ExcelWriter(iox,engine='openpyxl')
    writer.book = book
    dfx.to_excel(writer,name,index = None)
    writer.save()  
def AddTandx(sheet_name,iox):#将开式温度换算另作一列，计算失重率另作一列   
    df = pd.read_excel(iox, sheet_name = sheet_name)
    index = list(df.index)
    list_1 = [df.iloc[i,0]+273 for i in index]
    list_2 = [1/df.iloc[i,2] for i in index]
    df['1/T'] = list_2
    list_3 = [1-((df.iloc[0,1])-(df.iloc[i,1]))/((df.iloc[0,1])-(df.iloc[max(index),1])) for i in index]
    list_4 = [-(np.log(i)) for i in list_3]
    list_5 = [i/((df.iloc[j,1])*(df.iloc[j,1])) for(i,j) in zip(list_4,index)]
    list_6 = [np.log(i) for i in list_5]
    df['Y'] = list_5          
    book = load_workbook(iox)
    writer = pd.ExcelWriter(iox,engine='openpyxl')
    writer.book = book
    df.to_excel(writer, sheet_name,index = None)
    writer.save()     
def resize( w_box, h_box, pil_image):        
        w, h = pil_image.size #获取图像的原始大小          
        f1 = 1.0*w_box/w        
        f2 = 1.0*h_box/h       
        factor = min([f1, f2])     
        width = int(w*factor)
        height = int(h*factor)       
        return pil_image.resize((width, height), Image.ANTIALIAS)
def Select():
    io = askopenfilename(title='Select your file',initialdir='C:\\Windows')
    if io != '':
        tkinter.messagebox.showinfo( message = '文件名:'+io)
        window = tk.Toplevel()#实例化一个窗口
        window.title('活化能表征')
        window.geometry('600x350')
        w_box = 600
        h_box = 350
        canvas = tk.Canvas(window, bg='white', height=150, width=500)
        pil_image = Image.open(r'pig.gif')#获得合适大小的图片
        pil_image_resized = resize(150, 150,pil_image)
        tk_image = ImageTk.PhotoImage(pil_image_resized) 
        image = canvas.create_image(250, 0, anchor='n',image = tk_image)
        canvas.pack()#不能忘掉pack！！
        tk.Label(window, text='Ti:', font=('Arial', 14)).place(x=100, y=185)
        tk.Label(window, text='Tf:', font=('Arial', 14)).place(x=100, y=205)
        var_Ti = tk.IntVar()
        var_Tf = tk.IntVar()
        tk.Entry(window, textvariable = var_Ti , font=('Arial', 14)).place(x=180,y=185)
        tk.Entry(window, textvariable = var_Tf , font=('Arial', 14)).place(x=180,y=205)
        window.title('Friedman')
    else:
        tkinter.messagebox.showinfo( message = '您没有选择任何文件')
    def run_():
        Ti = var_Ti.get()   
        Tf = var_Tf.get()
        df = pd.read_excel(io, sheet_name =0,header=2)
        df1 = df.loc[(df["Temperature (°C)"]>Ti)&(df["Temperature (°C)"]<Tf)]
        Addsheet(df1,'df2',io)
        df = pd.read_excel(io,sheet_name = 1)
        index = list(df.index)
        list_1 = [df.iloc[i,0]+273 for i in index]
        df['T'] = list_1
        list_2 = [1/df.iloc[i,2] for i in index]
        df['1/T'] = list_2
        list_3 = [1-((df.iloc[0,1])-(df.iloc[i,1]))/((df.iloc[0,1])-(df.iloc[max(index),1]))for i in index]
        list_4 = [-(np.log(i)) for i in list_3]
        list_5 = [i/((df.iloc[j,2])*(df.iloc[j,2])) for(i,j) in zip(list_4,index)]
        list_6 = [np.log(i) for i in list_5]
        df['Y'] = list_6
        T_begin = Ti+273+20
        T_end = Tf+273-20
        df2 = df.loc[(df["T"]>T_begin)&(df["T"]<T_end)]
        parameter = np.polyfit(df2['1/T'],df2['Y'],1)
        x = df2["1/T"]
        y = parameter[0]*x+parameter[1]
        y_pred = [i*parameter[0]+parameter[1] for i in x]
        R_square =  r2_score(df2['Y'],y_pred)
        E = -8.314*parameter[0]/1000
        tkinter.messagebox.showinfo( message = f'R² = {R_square },E = {E}KJ ')
        book = load_workbook(io)
        writer = pd.ExcelWriter(io,engine='openpyxl')
        writer.book = book
        df.to_excel(writer,"Output",index=None)
        writer.save()   
        tkinter.messagebox.showinfo(title='运行窗口', message='计算完毕！ ')
        sheet_remove = ['df2','df21']
        for i in sheet_remove:
             book.remove_sheet(book.get_sheet_by_name(i))
        book.save(io)
        
    btn_Run = tk.Button(window, text='Run', command = run_ ).place(x=210, y=315)
    window.mainloop() 
window = tk.Tk()#实例化一个窗口
window.title('猪入文件~')
window.geometry('550x200')
w_box = 600
h_box = 350
canvas = tk.Canvas(window, bg='white', height=150, width=500)
pil_image = Image.open(r'pig2.gif')#获得合适大小的图片
pil_image_resized = resize(150, 150,pil_image)
tk_image = ImageTk.PhotoImage(pil_image_resized) 
image = canvas.create_image(250, 0, anchor='n',image = tk_image)
canvas.pack()
btn_Run = tk.Button(window, text='select your file', command = Select).place(x=210, y=160)
window.mainloop()


# In[ ]:





# In[ ]:




