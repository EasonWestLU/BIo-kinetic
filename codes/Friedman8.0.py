#!/usr/bin/env python
# coding: utf-8

# In[1]:


import tkinter as tk
import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
import io
from PIL import Image, ImageTk
import tkinter as tk
import tkinter.messagebox
from sklearn.metrics import r2_score
import matplotlib.pyplot as plt
import warnings
from tkinter.filedialog import askopenfilename
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.backend_bases import key_press_handler
from matplotlib.figure import Figure

def FindClosest(x,rate):#参数：一个数据list和一个目标rate的list    
    if rate>=x[-1]:        
        return x[-1]    
    elif rate<=x[0]:        
        return x[0]    
    x_less = []  
    for i in x:        
        if i<rate:           
            x_less.append(i)    
    x_more = []    
    for i in x:        
        if i>rate:            
            x_more.append(i)
                        
    a = max(x_less)    
    b = min(x_more)    
    if abs(rate - a) < abs(rate-b):        
        return a    
    elif abs(rate-b) < abs(rate-a):       
        return b    
    
def Addsheet(dfx,name,iox):#增加一个工作表，表名称        
    book = load_workbook(iox)
    writer=pd.ExcelWriter(iox,engine='openpyxl')
    writer.book = book
    dfx.to_excel(writer,name,index = None)
    writer.save()
    
def AddTandx(sheet_name,iox):#将开式温度换算另作一列，计算失重率另作一列   
    df = pd.read_excel(iox, sheet_name = sheet_name)
    index = list(df.index)
    list_1 = [df.iloc[i,0]+273 for i in index]
    df['T'] = list_1
    list_2 = [((df.iloc[0,1])-(df.iloc[i,1]))/((df.iloc[0,1])-(df.iloc[max(index),1])) for i in index]
    df['α'] = list_2
    
    def Sca_Deriv(x, y):       
        diff_x = [j-i for (i,j) in zip(x[0:], x[1:])]        
        diff_y = [j-i for (i,j) in zip(y[0:], y[1:])]      
        slopes = [diff_y[i]/diff_x[i] for i in range(len(diff_y))]        
        deriv = [0.5*(i+j) for (i,j) in zip(slopes[0:],slopes[1:])]        
        deriv.insert(0,slopes[0])        
        deriv.append(slopes[-1])        
        return deriv
    df['dα/dT'] = Sca_Deriv(list_1,list_2)                            
    book = load_workbook(iox)
    writer = pd.ExcelWriter(iox,engine='openpyxl')
    writer.book = book
    df.to_excel(writer, sheet_name,index = None)
    writer.save()   
    
def resize( w_box, h_box, pil_image):        
        w, h = pil_image.size #获取图像的原始大小          
        f1 = 1.0*w_box/w        
        f2 = 1.0*h_box/h       
        factor = min([f1, f2])     
        width = int(w*factor)
        height = int(h*factor)       
        return pil_image.resize((width, height), Image.ANTIALIAS)
def Select():
    io = askopenfilename(title='Select your file',initialdir='C:\\Windows')
    if io != '':
        tkinter.messagebox.showinfo( message = '文件名:'+io)
        window = tk.Toplevel()#实例化一个窗口
        window.title('活化能表征')
        window.geometry('600x350')
        w_box = 600
        h_box = 350
        canvas = tk.Canvas(window, bg='white', height=150, width=500)
        pil_image = Image.open(r'pig2.gif')#获得合适大小的图片
        pil_image_resized = resize(150, 150,pil_image)
        tk_image = ImageTk.PhotoImage(pil_image_resized) 
        image = canvas.create_image(250, 0, anchor='n',image = tk_image)
        canvas.pack()#不能忘掉pack！！
        tk.Label(window, text='T1i:', font=('Arial', 14)).place(x=100, y=185)
        tk.Label(window, text='T1f:', font=('Arial', 14)).place(x=100, y=205)
        tk.Label(window, text='T2i:', font=('Arial', 14)).place(x=100, y=225)
        tk.Label(window, text='T2f:', font=('Arial', 14)).place(x=100, y=245)
        tk.Label(window, text='T3i:', font=('Arial', 14)).place(x=100, y=265)
        tk.Label(window, text='T3f:', font=('Arial', 14)).place(x=100, y=285)
        var_T1i = tk.IntVar()
        var_T1f = tk.IntVar()
        var_T2i = tk.IntVar()
        var_T2f = tk.IntVar()
        var_T3i = tk.IntVar()
        var_T3f = tk.IntVar()
        tk.Entry(window, textvariable = var_T1i , font=('Arial', 14)).place(x=180,y=185)
        tk.Entry(window, textvariable = var_T1f , font=('Arial', 14)).place(x=180,y=205)
        tk.Entry(window, textvariable = var_T2i , font=('Arial', 14)).place(x=180,y=225)
        tk.Entry(window, textvariable = var_T2f , font=('Arial', 14)).place(x=180,y=245)
        tk.Entry(window, textvariable = var_T3i , font=('Arial', 14)).place(x=180,y=265)
        tk.Entry(window, textvariable = var_T3f , font=('Arial', 14)).place(x=180,y=285)   
        window.title('Friedman')
    else:
        tkinter.messagebox.showinfo( message = '您没有选择任何文件')
    def run_():
        T1i = var_T1i.get()   
        T1f = var_T1f.get()
        T2i = var_T2i.get()   
        T2f = var_T2f.get()
        T3i = var_T3i.get()   
        T3f = var_T3f.get()   
        class Cal:
            def __init__(self,a,b,c,d,e,f):
                self.sheet_name_1 = a
                self.min_ = b
                self.max_ = c
                self.sheet_name_2 = d
                self.sheet_name_3 = e
                self.Tratio = f
            def Run(self):               
                df = pd.read_excel(io, sheet_name = self.sheet_name_1,header=2)
                df1 = df.loc[(df["Temperature (°C)"]>self.min_)&(df["Temperature (°C)"]<self.max_ )]
                Addsheet(df1,self.sheet_name_2,io)
                AddTandx(self.sheet_name_2,io)
                df = pd.read_excel(io, sheet_name = self.sheet_name_3)
                list_α = list(df['α'])
                rate_appro = []
                rate = [0.1,0.2,0.3,0.4,0.5,0.6,0.7,0.8,0.9]
                for i in rate:
                    k = FindClosest(list_α, i)            
                    rate_appro.append(k)                    

                self.list_X = []        
                self.list_T = []             
                self.list_Re = []
                self.list_Y = [] 
                for i in rate_appro:                  
                    df_x = df[df['α'].isin([i])]   
                    j = 1/df_x.iloc[0,2]#1/T    
                    k = df_x.iloc[0,2]#T
                    l = df_x.iloc[0,4]
                    self.list_X.append(j)    
                    self.list_T.append(k)
                    self.list_Re.append(l)
                self.list_Y = [np.log(self.Tratio*i) for i in self.list_Re]
                dfr=pd.DataFrame()
                dfr['X']=self.list_X
                dfr['Y']=self.list_Y
                Addsheet(dfr,'Q',io)
        Data1 = Cal(0,T1i,T1f,'T1','T11',10)
        Data1.Run()
        Data2 = Cal(1,T2i,T2f,'T2','T21',20)
        Data2.Run()
        Data3 = Cal(2,T3i,T3f,'T3','T31',30)
        Data3.Run()
        
        root = tk.Tk()  # 创建tkinter的主窗口
        name = ['01','02','03','04','05','06','07','08','09']
        rate = [0.1,0.2,0.3,0.4,0.5,0.6,0.7,0.8,0.9]
        f = plt.figure(figsize = (13,10),dpi=180)
        a = f.add_subplot(111) 
        for i,j in zip(name,np.arange(0,10)):
            exec('X_sca%s = [Data1.list_X[j],Data2.list_X[j],Data3.list_X[j]]'%i)         
            exec('Y_sca%s = [Data1.list_Y[j],Data2.list_Y[j],Data3.list_Y[j]]'%i)
        a.scatter(Data1.list_X,Data1.list_Y,label = '10K/min',marker = 'o')
        a.scatter(Data2.list_X,Data2.list_Y,label = '20K/min',marker = 'v')
        a.scatter(Data3.list_X,Data3.list_Y,label = '30K/min',marker = '*')
        for i in name:    
            exec('par%s = np.polyfit(X_sca%s,Y_sca%s,1)'%(i,i,i))
        for i in name:    
            exec('x%s = np.array(X_sca%s)'%(i,i))
        for i in name:    
            exec('Line%s = par%s[0]*x%s + par%s[1]'%(i,i,i,i))
        s = '--'
        for i in name:    
            exec('a.plot(X_sca%s,Line%s,linestyle= s)'%(i,i))
        list_Ea  = []
        for i in name:       
            exec('list_Ea.append((-8.314*par%s[0])/1000)'%i)    
        Ea_mean = np.mean(list_Ea)
        R_square = []
        for i in name:
            exec('R_square%s = r2_score(Y_sca%s,Line%s)'%(i,i,i))
        for i in name:
            exec('R_square.append(R_square%s)'%i)
        Ea_α = {'α(%)':rate,'Ea(KJ/mol)':list_Ea,'Mean':Ea_mean,'R²':R_square}
        data = pd.DataFrame(Ea_α)
        Ea = 'Ea'
        a.set_xlabel('1/T', fontproperties = 'Times New Roman',  size = 22,  fontweight = 'bold')
        a.set_ylabel('ln(β*(dα/dT))', fontproperties = 'Times New Roman', size = 22, fontweight = 'bold')
        canvas = FigureCanvasTkAgg(f, master=root)
        canvas.draw()
        canvas.get_tk_widget().pack(side=tkinter.TOP,  # 上对齐
                            fill=tkinter.BOTH,  # 填充方式
                            expand=tkinter.YES)

        tkinter.messagebox.showinfo( message = f'活化能均值:{Ea_mean}')
        
        root1 = tk.Tk()
        root1.title('判断窗口')
        root1.geometry('300x200')
        def Save():
            Addsheet(data,Ea,io)
            sheet_remove = ['T1','T2','T3']
            book = load_workbook(io)
            sheet_newName = ['Res1','Res2','Res3']
            sheet_OldName = ['T11','T21','T31']
            for i,j in zip(sheet_OldName,sheet_newName):    
                ws = book[i]    
                ws.title = j                
            for i in sheet_remove:    
                book.remove_sheet(book.get_sheet_by_name(i))    
                book.save(io)
        def Clear():
            book = load_workbook(io)
            sheet_remove = ['T1','T2','T3','T11','T21','T31','Q','Q1','Q2']
            for i in sheet_remove:
                book.remove_sheet(book.get_sheet_by_name(i))    
                book.save(io)
            
            
           
        tk.Label(root1,text='是否保存数据?', font=('Arial', 10)).place(x=100, y=100)
               
        btn_Run = tk.Button( root1, text='是', command = Save).place(x=80, y=160)
        btn_Run = tk.Button( root1, text='否', command = Clear).place(x=180, y=160)
        root1.mainloop()
        
        root.mainloop()               
    btn_Run = tk.Button(window, text='Run', command = run_ ).place(x=210, y=315)
    window.mainloop() 
window = tk.Tk()#实例化一个窗口
window.title('猪入文件~')
window.geometry('550x200')
w_box = 600
h_box = 350
canvas = tk.Canvas(window, bg='white', height=150, width=500)
pil_image = Image.open(r'pig.gif')#获得合适大小的图片
pil_image_resized = resize(150, 150,pil_image)
tk_image = ImageTk.PhotoImage(pil_image_resized) 
image = canvas.create_image(250, 0, anchor='n',image = tk_image)
canvas.pack()
btn_Run = tk.Button(window, text='select your file', command = Select).place(x=210, y=160)
window.mainloop()


# In[14]:


#绘图
import pandas as pd
import numpy as np
import os
from sklearn.metrics import r2_score
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
io = 'C:\\Users\\Wang Pu\\Desktop\\Straw-Do-F.xlsx'
df1 = pd.read_excel(io, sheet_name = 'Q')
df2 = pd.read_excel(io, sheet_name = 'Q1')
df3 = pd.read_excel(io, sheet_name = 'Q2')
X1 = list(df1['X'])
X2 = list(df2['X'])
X3 = list(df3['X'])
Y1 = list(df1['Y'])
Y2 = list(df2['Y'])
Y3 = list(df3['Y'])
name = ['01','02','03','04','05','06','07','08','09']
for i,j in zip(name,np.arange(0,10)):    
    exec('X_sca%s = [X1[j],X2[j],X3[j]]'%i)         
    exec('Y_sca%s = [Y1[j],Y2[j],Y3[j]]'%i)     
for i in name:    
    exec('par%s = np.polyfit(X_sca%s,Y_sca%s,1)'%(i,i,i))
for i in name:    
    exec('x%s = np.array(X_sca%s)'%(i,i))
for i in name:    
    exec('Line%s = par%s[0]*x%s + par%s[1]'%(i,i,i,i))

plt.figure(figsize = (13,10),dpi=180)
plt.scatter(X1,Y1,label = '10K/min',marker = 'o')
plt.scatter(X2,Y2,label = '20K/min',marker = 'v')
plt.scatter(X3,Y3,label = '30K/min',marker = '*')
s = '--'
for i in name:    
    exec('plt.plot(X_sca%s,Line%s,linestyle= s)'%(i,i))
plt.xticks(fontproperties = 'Times New Roman', size = 18, weight = 'bold')
plt.yticks(fontproperties = 'Times New Roman', size = 18, weight = 'bold')
plt.xlabel('1/T', fontproperties = 'Times New Roman',  size = 22,  fontweight = 'bold')
plt.ylabel('ln(β*(dα/dT))', fontproperties = 'Times New Roman', size = 22, fontweight = 'bold')
ax = plt.gca()
ax.spines['bottom'].set_linewidth(2.5)
ax.spines['left'].set_linewidth(2.5)
ax.spines['top'].set_linewidth(2.5)
ax.spines['right'].set_linewidth(2.5)
plt.text(0.00185,-3,'α = 0.1', fontproperties = 'Times New Roman',  size = 22)
plt.text(0.00160,-3.1,'α = 0.9', fontproperties = 'Times New Roman',  size = 22)
plt.text(0.00176,-1,'(Straw+Calcined Do)', fontproperties = 'Times New Roman',  size = 30)


# In[ ]:




